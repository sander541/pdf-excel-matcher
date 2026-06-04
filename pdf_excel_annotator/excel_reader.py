from __future__  import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .utils import normalize_code


@dataclass
class ExcelCodeEntry:

    code_raw: str
    code_norm: str
    excel_row: int
    row_data: Sequence[Tuple[str, str]]
    expected_count: int = 1  # From count_column, defaults to 1 if not specified
    specifier_norm: str | None = None  # Normalized value from specifier_column


def read_column_headers(
    workbook_path: str | Path,
    header_row: int,
) -> List[Tuple[str, str]]:
    """Return [(column_letter, header_name), ...] for every column in the header row."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        if not rows:
            return []
        result: List[Tuple[str, str]] = []
        for i, v in enumerate(rows[0]):
            letter = get_column_letter(i + 1)
            name = str(v).strip() if v is not None and str(v).strip() else letter
            result.append((letter, name))
        return result
    finally:
        workbook.close()


def load_excel_codes(
    workbook_path: str | Path,
    code_column: str,
    header_row: int,
    max_row: int | None = None,
    count_column: str | None = None,
    specifier_column: str | None = None,
    annotation_columns: Optional[frozenset] = None,
) -> List[ExcelCodeEntry]:

    # Convert column letters to 0-based indices for tuple access.
    code_col_idx = column_index_from_string(code_column.upper()) - 1
    count_col_idx = (column_index_from_string(count_column.upper()) - 1) if count_column else None
    spec_col_idx = (column_index_from_string(specifier_column.upper()) - 1) if specifier_column else None

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        final_row = sheet.max_row if max_row is None else min(max_row, sheet.max_row)

        # Read all needed rows in a single sequential scan — avoids the per-cell
        # XML re-parse that makes random sheet.cell() access O(rows²) in read-only mode.
        # Trade-off: the entire range is materialised into memory at once. For typical
        # workbooks (hundreds of rows, tens of columns) this is negligible. For very
        # large sheets with max_row=None, consider passing an explicit max_row cap.
        all_rows = list(
            sheet.iter_rows(min_row=header_row, max_row=final_row, values_only=True)
        )
        if not all_rows:
            return []

        # Build header labels from the first row.
        header_vals = all_rows[0]
        headers: list[str] = [
            (str(v).strip() if v is not None and str(v).strip() else get_column_letter(i + 1))
            for i, v in enumerate(header_vals)
        ]

        entries: List[ExcelCodeEntry] = []
        for offset, row_vals in enumerate(all_rows[1:], start=1):
            row_idx = header_row + offset

            if code_col_idx >= len(row_vals):
                continue
            value = row_vals[code_col_idx]
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            normalized = normalize_code(text)
            if not normalized:
                continue

            # Determine which 0-based column indices to include in annotation details.
            # annotation_columns (set of letters) → include exactly those columns.
            # No annotation_columns → include all columns except the count column.
            if annotation_columns is not None:
                include_indices: set[int] | None = {
                    column_index_from_string(col.upper()) - 1
                    for col in annotation_columns
                }
            else:
                include_indices = None  # all columns; count excluded below

            row_entries: list[Tuple[str, str]] = []
            for ci, cell_val in enumerate(row_vals):
                if cell_val is None:
                    continue
                val_text = str(cell_val).strip()
                if not val_text:
                    continue
                if include_indices is not None:
                    if ci not in include_indices:
                        continue
                else:
                    # Default: exclude count column so it doesn't clutter popups
                    if count_col_idx is not None and ci == count_col_idx:
                        continue
                header = headers[ci] if ci < len(headers) else get_column_letter(ci + 1)
                row_entries.append((header, val_text))

            expected_count = 1
            if count_col_idx is not None and count_col_idx < len(row_vals):
                cv = row_vals[count_col_idx]
                if cv is not None:
                    try:
                        expected_count = max(1, int(cv))
                    except (ValueError, TypeError):
                        pass

            specifier_norm: Optional[str] = None
            if spec_col_idx is not None and spec_col_idx < len(row_vals):
                sv = row_vals[spec_col_idx]
                if sv is not None:
                    spec_text = str(sv).strip()
                    specifier_norm = normalize_code(spec_text) or None

            entries.append(
                ExcelCodeEntry(
                    code_raw=text,
                    code_norm=normalized,
                    excel_row=row_idx,
                    row_data=row_entries,
                    expected_count=expected_count,
                    specifier_norm=specifier_norm,
                )
            )
        return entries
    finally:
        workbook.close()
